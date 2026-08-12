"""
Vietnamese Accounting Software (MISA / FAST) & VAT Schedule (Bảng Kê 01-1/GTGT) Exporter.
Generates official tax schedule reports and 1-click MISA SME / AMIS import workbooks.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("nesy_docai.misa_vat_exporter")


class MISAandVATScheduleExporter:
    """
    Exports parsed & verified invoice audit records into:
    1. Bảng kê thuế GTGT mua vào (Form 01-1/GTGT per Vietnamese General Dept of Taxation standards)
    2. MISA SME / AMIS Excel Import Template (Direct 1-click accounting software import)
    """

    def export_vat_schedule_excel(
        self,
        records: List[Dict[str, Any]],
        output_filepath: Union[str, Path] = "Bang_Ke_Thue_GTGT_01_1.xlsx",
        company_name: str = "ĐƠN VỊ SỬ DỤNG",
        company_mst: str = "0100000000"
    ) -> Path:
        output_path = Path(output_filepath)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Styles
        font_title = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        font_subtitle = Font(name="Segoe UI", size=10, italic=True)
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        border_thick_bottom = Border(bottom=Side(style='medium', color='1F4E78'))

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # ------------------------------------------------------------------
        # SHEET 1: BẢNG KÊ HÓA ĐƠN MUA VÀO (MẪU 01-1/GTGT)
        # ------------------------------------------------------------------
        ws_vat = wb.active
        ws_vat.title = "Bảng Kê GTGT 01-1 Mua Vào"

        ws_vat.append([f"BẢNG KÊ HÓA ĐƠN, CHỨNG TỪ HÀNG HÓA, DỊCH VỤ MUA VÀO"])
        ws_vat.append([f"Tên người nộp thuế: {company_name} | MST: {company_mst}"])
        ws_vat.append(["(Kèm theo Tờ khai thuế GTGT)"])
        ws_vat.append([])  # blank row

        ws_vat.cell(row=1, column=1).font = font_title
        ws_vat.cell(row=2, column=1).font = font_subtitle
        ws_vat.cell(row=3, column=1).font = font_subtitle

        headers = [
            "STT",
            "Ký Hiệu Mẫu Số",
            "Ký Hiệu Hóa Đơn",
            "Số Hóa Đơn",
            "Ngày Lập",
            "Tên Người Bán",
            "Mã Số Thuế Người Bán",
            "Nội Dung Hàng Hóa Dịch Vụ Mua Vào",
            "Doanh Số Mua Chưa Thuế (VNĐ)",
            "Thuế Suất GTGT (%)",
            "Tiền Thuế GTGT (VNĐ)",
            "Mã Thuế Suất",
            "Trạng Thái Kiểm Toán (Z3 & MST)"
        ]

        ws_vat.append(headers)
        header_row_idx = 5

        for col_idx in range(1, len(headers) + 1):
            cell = ws_vat.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        ws_vat.row_dimensions[header_row_idx].height = 28

        # Group records by tax rate
        total_subtotal_all = 0
        total_tax_all = 0

        for idx, rec in enumerate(records, start=1):
            subtotal_val = self._parse_int(rec.get("subtotal"))
            tax_val = self._parse_int(rec.get("tax"))
            tax_rate_str = str(rec.get("tax_rate", "10%"))

            total_subtotal_all += subtotal_val
            total_tax_all += tax_val

            # Determine vendor tax code (0, 5, 8, 10)
            clean_rate_num = tax_rate_str.replace("%", "").strip()
            vendor_tax_code = clean_rate_num if clean_rate_num in ["0", "5", "8", "10"] else "10"

            status_badge = rec.get("audit_status", "SAT")
            if rec.get("is_xml"):
                status_badge = "✅ XÁC THỰC XML GỐC"
            elif "SAT" in status_badge:
                status_badge = "✅ VERIFIED Z3 Presburger"
            else:
                status_badge = f"⚠️ {status_badge}"

            row = [
                idx,
                rec.get("invoice_form", "1"),
                rec.get("invoice_symbol", rec.get("invoice_series", "")),
                rec.get("invoice_number", rec.get("invoice_id", "")),
                rec.get("invoice_date", ""),
                rec.get("seller_name", ""),
                rec.get("seller_tax_id", ""),
                rec.get("description_summary", ""),
                subtotal_val,
                tax_rate_str,
                tax_val,
                vendor_tax_code,
                status_badge
            ]

            ws_vat.append(row)
            curr_row = ws_vat.max_row
            
            ws_vat.cell(row=curr_row, column=1).alignment = center_align
            ws_vat.cell(row=curr_row, column=2).alignment = center_align
            ws_vat.cell(row=curr_row, column=3).alignment = center_align
            ws_vat.cell(row=curr_row, column=4).alignment = center_align
            ws_vat.cell(row=curr_row, column=5).alignment = center_align
            ws_vat.cell(row=curr_row, column=6).alignment = left_align
            ws_vat.cell(row=curr_row, column=7).alignment = center_align
            ws_vat.cell(row=curr_row, column=8).alignment = left_align
            ws_vat.cell(row=curr_row, column=9).alignment = right_align
            ws_vat.cell(row=curr_row, column=9).number_format = '#,##0'
            ws_vat.cell(row=curr_row, column=10).alignment = center_align
            ws_vat.cell(row=curr_row, column=11).alignment = right_align
            ws_vat.cell(row=curr_row, column=11).number_format = '#,##0'
            ws_vat.cell(row=curr_row, column=12).alignment = center_align
            ws_vat.cell(row=curr_row, column=13).alignment = left_align

        # Total Summary Row
        summary_row = [
            "TỔNG CỘNG", "", "", "", "", "", "", "Tổng doanh số & thuế GTGT mua vào",
            total_subtotal_all, "", total_tax_all, "", ""
        ]
        ws_vat.append(summary_row)
        sum_row_idx = ws_vat.max_row

        for col_idx in range(1, len(headers) + 1):
            cell = ws_vat.cell(row=sum_row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10, bold=True)
            cell.fill = summary_fill
            if col_idx in [9, 11]:
                cell.alignment = right_align
                cell.number_format = '#,##0'

        # ------------------------------------------------------------------
        # SHEET 2: MISA SME / AMIS IMPORT TEMPLATE
        # ------------------------------------------------------------------
        ws_misa = wb.create_sheet(title="MISA Import Mua Vào")

        misa_headers = [
            "Ngày hạch toán", "Ngày chứng từ", "Số chứng từ", "Số hóa đơn",
            "Mã đối tượng (MST)", "Tên đối tượng (Người bán)", "Địa chỉ",
            "Diễn giải", "Mã hàng", "Tên hàng", "TK Chi phí / Mua hàng",
            "Số lượng", "Đơn giá", "Thành tiền", "Thuế suất", "Tiền thuế GTGT"
        ]

        ws_misa.append(misa_headers)
        for col_idx in range(1, len(misa_headers) + 1):
            cell = ws_misa.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for rec in records:
            inv_date = rec.get("invoice_date", "")
            inv_no = rec.get("invoice_number", rec.get("invoice_id", ""))
            mst = rec.get("seller_tax_id", "")
            seller = rec.get("seller_name", "")
            address = rec.get("seller_address", "")
            desc = rec.get("description_summary", "")

            # If multi line items exist
            line_items = rec.get("line_items", [])
            if line_items and isinstance(line_items, list):
                for item in line_items:
                    misa_row = [
                        inv_date, inv_date, f"CT-{inv_no}", inv_no,
                        mst, seller, address,
                        item.get("description", desc),
                        f"VT-{item.get('item_id', 1)}", item.get("description", desc),
                        "6422",  # Default accounting cost account
                        self._parse_int(item.get("quantity", 1)),
                        self._parse_int(item.get("unit_price", 0)),
                        self._parse_int(item.get("amount", rec.get("subtotal", 0))),
                        rec.get("tax_rate", "10%"),
                        self._parse_int(rec.get("tax", 0))
                    ]
                    ws_misa.append(misa_row)
            else:
                misa_row = [
                    inv_date, inv_date, f"CT-{inv_no}", inv_no,
                    mst, seller, address, desc, "DV-01", desc, "6422",
                    1, self._parse_int(rec.get("subtotal", 0)),
                    self._parse_int(rec.get("subtotal", 0)),
                    rec.get("tax_rate", "10%"),
                    self._parse_int(rec.get("tax", 0))
                ]
                ws_misa.append(misa_row)

        # Apply borders and auto-fit column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row > 4:
                        cell.border = border_thin
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        logger.info("Successfully exported VAT Schedule & MISA import file to: %s", output_path)
        return output_path

    @staticmethod
    def _parse_int(val: Any) -> int:
        if val is None or val == "" or val == "N/A":
            return 0
        try:
            cleaned = str(val).replace(",", "").replace(".", "").replace("đ", "").strip()
            return int(float(cleaned))
        except (ValueError, TypeError):
            return 0
