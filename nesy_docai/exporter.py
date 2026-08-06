"""
Audit Excel Exporter Module (Multi-Sheet Audit Ledger & Provenance Exporter)
"""

import json
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class AuditExcelExporter:
    def export_to_excel(self, audit_records: List[Dict[str, Any]], output_filepath: str = "invoice_audit_report.xlsx") -> str:
        """
        Exports verified invoice audit data into a formatted multi-sheet Excel workbook.
        """
        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        sat_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        unsat_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # -------------------------------------------------------------
        # SHEET 1: Summary (Bảng Tổng Hợp Hóa Đơn)
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Invoice Summary"

        summary_headers = [
            "Invoice ID", "Date", "Seller Tax ID", "Seller Name",
            "Subtotal (VND)", "Tax (VND)", "Total (VND)",
            "Audit Status", "Tax Verification"
        ]

        ws_summary.append(summary_headers)
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in audit_records:
            tax_info = record.get("tax_verification", {})
            tax_status_str = tax_info.get("status", "N/A") if isinstance(tax_info, dict) else "N/A"

            row_data = [
                record.get("invoice_id"),
                record.get("invoice_date"),
                record.get("seller_tax_id"),
                record.get("seller_name"),
                record.get("subtotal"),
                record.get("tax"),
                record.get("total"),
                record.get("audit_status"),
                tax_status_str
            ]
            ws_summary.append(row_data)

            current_row = ws_summary.max_row
            status_cell = ws_summary.cell(row=current_row, column=8)
            if record.get("audit_status") == "VERIFIED_SAT":
                status_cell.fill = sat_fill
            else:
                status_cell.fill = unsat_fill

        # -------------------------------------------------------------
        # SHEET 2: Line Items Detail (Chi Tiết Mặt Hàng)
        # -------------------------------------------------------------
        ws_items = wb.create_sheet(title="Line Items Detail")
        item_headers = [
            "Invoice ID", "Item #", "Description", "Quantity",
            "Unit Price (VND)", "Amount (VND)", "Bounding Box [x0, y0, x1, y1]"
        ]
        ws_items.append(item_headers)
        for col_idx in range(1, len(item_headers) + 1):
            cell = ws_items.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in audit_records:
            inv_id = record.get("invoice_id")
            for item in record.get("line_items", []):
                bbox_str = str(item.get("bbox", []))
                ws_items.append([
                    inv_id,
                    item.get("item_id"),
                    item.get("description"),
                    item.get("quantity"),
                    item.get("unit_price"),
                    item.get("amount"),
                    bbox_str
                ])

        # -------------------------------------------------------------
        # SHEET 3: Audit Ledger Log (Bằng Chứng Kiểm Toán Z3)
        # -------------------------------------------------------------
        ws_audit = wb.create_sheet(title="Audit Trail Log")
        audit_headers = ["Invoice ID", "SMT Status", "Proof Certificate", "Constraints Verified"]
        ws_audit.append(audit_headers)
        for col_idx in range(1, len(audit_headers) + 1):
            cell = ws_audit.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in audit_records:
            cert = record.get("proof_certificate", {})
            constraints = ", ".join(cert.get("constraints_verified", []))
            ws_audit.append([
                record.get("invoice_id"),
                cert.get("smt_status"),
                cert.get("proof_formula", cert.get("reason", "")),
                constraints
            ])

        # Auto-fit column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.border = border_thin
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_filepath)
        return output_filepath
