"""
Audit Excel Ledger Report Exporter
"""

import os
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class AuditExcelExporter:
    def export_to_excel(self, records: List[Dict[str, Any]], output_filepath: str = "invoice_audit_report.xlsx") -> str:
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary Audit Ledger"

        # Headers
        headers = [
            "File Name", "Invoice ID", "Date", "Vendor Tax Code",
            "Vendor Name", "Description Summary", "Subtotal (VND)",
            "Tax Rate", "Tax Amount (VND)", "Total (VND)",
            "Audit Status", "Z3 SMT Proof Certificate"
        ]

        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        ws_summary.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_summary.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for r in records:
            status = r.get("audit_status", "UNKNOWN")
            is_sat = (status == "VERIFIED_SAT" or status == "SAT")

            row = [
                r.get("file_name", r.get("invoice_id", "N/A")),
                r.get("invoice_id", "N/A"),
                r.get("invoice_date", "N/A"),
                r.get("seller_tax_id", "N/A"),
                r.get("seller_name", "N/A"),
                r.get("description_summary", "N/A"),
                r.get("subtotal", 0),
                r.get("tax_rate", "0%"),
                r.get("tax", 0),
                r.get("total", 0),
                "✅ VERIFIED_SAT" if is_sat else f"❌ {status}",
                r.get("certificate", "Z3 Presburger Bounds Verified") if is_sat else "UNSAT Constraint Failure"
            ]
            ws_summary.append(row)

        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_filepath)
        return os.path.abspath(output_filepath)
